import openpyxl as pxl  #To handle a excel file
import random as rand   #To generate random number
import generate_Latex_Format as latex 
import JSON_Formatters 
import json
import print_Question_Options as display
'''
Author Name : Bhathrinaathan M B
Description : Generates a name the thing or place or the person typr questions using the data from excel files
Date        : 1st August 2022
Last updated on : 3rd August 2022
'''
author_Name='Bhathrinaathan M B'
guid_Id='08d3ea48-d909-4d98-8c45-81d6abaceb1b'
sys_Id='Hexint09'
reveiwer_Name='Tagore'
reveiwer_Id=''

wb=pxl.load_workbook("/home/bhathri/HexFace/Name It.xlsx")    
sheet=wb.active
Size=sheet.max_row #To find the maximum size of the sheet


#Generates options for given number of count , selects data from the specified coloumn(argument col) in excel sheet
def generate_options(row,col,number_of_options,boundry):
    i=rand.randint(boundry[0],boundry[1])
    options=[sheet.cell(row,col).value] #Appends the answer to the list
    while len(options)<number_of_options:
        if sheet.cell(i,5).value not in options:
            options.append(sheet.cell(i,col).value)
        i+=1
    rand.shuffle(options)
    return options

#------------------------------------------WHAT?----------------------------------------------------#

def  generate_what_type_ques(row,number_of_options):

    ques,col=rand.choice([generate_detail_by_what(row),generate_what_by_details(row)]) #Randomly chooses a question format
    options=generate_options(row,col,number_of_options,(60,170)) #The limit for the question types
    ques_latex,options_latex,answer_latex=latex.format_name_it(ques,options,sheet.cell(row,col).value)
    data=JSON_Formatters.format_json_file(ques_latex,options_latex,answer_latex,author_Name,guid_Id,sys_Id,reveiwer_Name,reveiwer_Id)
    json.dumps(data)
    print('_'*120)
    display.print_questions_option_answer(ques,options,sheet.cell(row,col).value)
    print(data)

def generate_what_by_details(row):
    ques=f"Which term is referred as {sheet.cell(row,4).value}"
    return ques,5 #5 is the col which is used for options
    

def generate_detail_by_what(row):
    ques=f"Choose the correct option for the term {sheet.cell(row,5).value}"
    return ques,4 #4th column is used to create options


#------------------------------------------WHO?----------------------------------------------------#


def generate_who_type_ques(row,number_of_options):
    ques,col=rand.choice([generate_detail_by_who(row),generate_who_by_details(row)])
    options=generate_options(row,col,number_of_options,(194,Size)) #The limit for the question types is 194 to last column
    ques_latex,options_latex,answer_latex=latex.format_name_it(ques,options,sheet.cell(row,col).value)
    data=JSON_Formatters.format_json_file(ques_latex,options_latex,answer_latex,author_Name,guid_Id,sys_Id,reveiwer_Name,reveiwer_Id)
    json.dumps(data)
    print('_'*120)
    display.print_questions_option_answer(ques,options,sheet.cell(row,col).value)
    print(data)

def generate_who_by_details(row):
    ques=f"Who is the {sheet.cell(row,4).value}"
    return ques,5 #5 is the col which is used for options
    

def generate_detail_by_who(row):
    ques=f"Choose the correct option for the following person {sheet.cell(row,5).value}"
    return ques,4 #4th column is used to create options



#------------------------------------------Where?----------------------------------------------------#

#Generate 'Where' type of questions
def generate_where_type_ques(row,number_of_options):
    ques1=f"Where is the {sheet.cell(row,4).value} located in India? "
    ques2=f"In which part of India , {sheet.cell(row,4).value} is located ? "
    ques=rand.choice([ques1,ques2])
    options=generate_options(row,5,number_of_options,(1,40))   #(2,40) is the limit where the required type of ques is found
    ques_latex,options_latex,answer_latex=latex.format_name_it(ques,options,sheet.cell(row,5).value)
    data=JSON_Formatters.format_json_file(ques_latex,options_latex,answer_latex,author_Name,guid_Id,sys_Id,reveiwer_Name,reveiwer_Id)
    json.dumps(data)
    print('_'*120)
    display.print_questions_option_answer(ques,options,sheet.cell(row,5).value)
    print(data)


#Generates a random question
def generate_random_question(number_of_options):
    row=rand.randint(1,Size)
    type=sheet.cell(row,2)
    ques_dict={'Where':generate_where_type_ques(row,number_of_options) , 
                'What':generate_what_type_ques(row,number_of_options),
                'Who' :generate_who_type_ques(row,number_of_options)}
    return ques_dict[type]
    
if __name__=='__main__':
    number_of_options=int(input("Enter the number of options required : "))
    generate_random_question(number_of_options) #A random question among any type
    generate_where_type_ques(rand.randint(2,58),number_of_options) #Row 2 to Row 58 have 'where' type data
    generate_what_type_ques(rand.randint(59,193),number_of_options) #Row 59 to Row 193 have 'what' type data
    generate_who_type_ques(rand.randint(194,Size),number_of_options) #Row 194 to last have 'who' type data